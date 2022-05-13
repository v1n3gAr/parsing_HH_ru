import requests
from openpyxl import Workbook
from datetime import datetime
from fpdf import FPDF
import os

# todo
# Собственный обработчик АПИ значений, к единому формату для всех классов - complete
# Аргументы командной строки и класс, приводящий отправку к единому формату, паттерн: Стратегия, возможно, Фабрика
# Нормальное описание и требование к вакансии, без троеточия...
# Поглядеть в целом про паттерны проектирования на Питоне

folder = os.path.dirname(os.path.abspath(__file__))


class Parser:
    data = {
            'name': [],
            'salary_to': [],
            'salary_from': [],
            'currency': [],
            'snippet_requirement': [],
            'snippet_responsibility': [],
            'alternate_url': []
        }

    def __init__(self, text: str):
        self.text = text  # Запрос, который мы будем передавать для поиска на hh.ru
        self._area = 66  # Код города Нижний Новгород, все id городов находятся в документации к API
        self._per_page = 100  # Количество отображаемых результатов
        self._url = 'https://api.hh.ru/vacancies'

    def __parsing(self):
        req = requests.get(self._url, params={
            'text': self.text,
            'area': self._area,
            'per_page': self._per_page
        }
                           )
        data = req.json()['items']
        req.close()
        return data

    @classmethod
    def __writing_to_the_internal_dictionary(cls, parsing):
        for row_number in range(len(parsing)):
            cls.data['name'].append(parsing[row_number]['name'])
            if parsing[row_number]["salary"] is not None:
                if parsing[row_number]["salary"]["to"] is not None:
                    cls.data['salary_from'].append(parsing[row_number]["salary"]["from"])
                    cls.data['salary_to'].append(parsing[row_number]["salary"]["to"])
                    cls.data['currency'].append(parsing[row_number]["salary"]["currency"])
                else:
                    cls.data['salary_from'].append(parsing[row_number]["salary"]["from"])
                    cls.data['currency'].append(parsing[row_number]["salary"]["currency"])
                    cls.data['salary_to'].append('Ограничение по зарплате не указано')
            else:
                cls.data['salary_from'].append(r'з\п не указана')
                cls.data['salary_to'].append(r'з\п не указана')
                cls.data['currency'].append(r'')

            cls.data['snippet_requirement'].append(
                parsing[row_number]['snippet']['requirement'].replace("<highlighttext>", "").replace("</highlighttext>",
                                                                                                     ""))
            if parsing[row_number]['snippet']['responsibility'] is not None:
                cls.data['snippet_responsibility'].append(
                    parsing[row_number]['snippet']['responsibility'].replace("<highlighttext>", "").replace(
                        "</highlighttext>", ""))
            else:
                cls.data['snippet_responsibility'].append(parsing[row_number]['snippet']['responsibility'])
            cls.data['alternate_url'].append(parsing[row_number]['alternate_url'])

        return cls.data

    def all_data_from_hh(self):
        print('Воспользовались интерфейсом для захвата данных с парсинга')
        return self.__writing_to_the_internal_dictionary(self.__parsing())


class Excel:
    def __init__(self, date_and_time):
        self.date_and_time = date_and_time  # Берем дату и время для их добавления к наименованию файла,
        # чтобы каждый новый файл был уникален
        self.wb = Workbook()
        self.ws = self.wb.active
        self.file_name = f"Парсинг за {self.date_and_time}.xlsx"
        headings = {
            'A1': 'Наименование вакансии',
            'B1': 'Зарплата',
            'C1': 'Требования',
            'D1': 'Обязанности',
            'E1': 'Ссылка',
        }

        for i, j in headings.items():  # Заполняем шапку Excel
            self.ws[i].value = j

    def __filling_in_data(self, cell, value):
        self.ws[cell].value = value

    def __save(self):
        self.wb.save(self.file_name)

    def __close(self):
        self.wb.close()

    def create_file(self, data_from_parsing):
        for row_number in range(len(data_from_parsing['name'])):
            self.__filling_in_data(f'A{row_number + 2}', data_from_parsing['name'][row_number])
            self.__filling_in_data(f'B{row_number + 2}', f'{data_from_parsing["salary_from"][row_number]} - {data_from_parsing["salary_to"][row_number]} {data_from_parsing["currency"][row_number]}')
            self.__filling_in_data(f'C{row_number + 2}', data_from_parsing['snippet_requirement'][row_number])
            self.__filling_in_data(f'D{row_number + 2}', data_from_parsing['snippet_responsibility'][row_number])
            self.__filling_in_data(f'E{row_number + 2}', data_from_parsing['alternate_url'][row_number])
        self.__close()
        self.__save()

        return self.file_name


class Pdf:
    def __init__(self, date_and_time):
        self.pdf = FPDF()
        self.w = 0
        self.h = 10
        self.ln = 2
        self.date_and_time = date_and_time
        self.file_name = f'Парсинг за {self.date_and_time}.pdf'

    def __create_file_pdf(self):
        self.pdf.add_page()

    def __create_text_on_page(self, text, align='J'):
        self.pdf.add_font("NotoSans", style="", fname="font/NotoSans-Bold.ttf", uni=True)
        self.pdf.set_font('NotoSans', size=14)
        self.pdf.multi_cell(self.w, self.h, txt=text, align=align)

    def __save_pdf(self):
        self.pdf.output(self.file_name)

    def create_file(self, data_from_parsing):
        self.__create_file_pdf()
        for i in range(len(data_from_parsing['name'])):
            self.__create_text_on_page('-' * 118)
            self.__create_text_on_page(data_from_parsing['name'][i], align='C')
            if data_from_parsing["snippet_requirement"][i] is not None and data_from_parsing["snippet_responsibility"][i] is not None:
                self.__create_text_on_page(f'Требования: {data_from_parsing["snippet_requirement"][i]}')
                self.__create_text_on_page(f'Описание: {data_from_parsing["snippet_responsibility"][i]}')
            else:
                self.__create_text_on_page(f'Описание или требование было неизвестно')
            if data_from_parsing["salary_from"][i] is not None and data_from_parsing["salary_to"][i] is not None:
                self.__create_text_on_page(
                    f'Зарплата: От - {data_from_parsing["salary_from"][i]} до {data_from_parsing["salary_to"][i]}')
            else:
                self.__create_text_on_page(f'Зарплата неизвестна')
        self.__save_pdf()

        return self.file_name

#todo:
#Передавать именно экземпляр класса, а не сам класс
#Передавать выбор стратегии через cmd


class FilesCreator:
    def __init__(self, strategy, data_from_parsing):
        now = datetime.now()
        self.date_and_time = now.strftime("%d.%m.%y %H_%M")
        self.strategy = strategy
        self.data_from_parsing = data_from_parsing

    def create_file(self):
        document = self.strategy(self.date_and_time)
        file_name = document.create_file(self.data_from_parsing)
        return file_name


class Telegram:
    data_for_send_msg = {
            'caption': 'Результат парсинга',
            'chat_id': '946919713',
            'url': 'https://api.telegram.org/bot5239048826:AAHYeNnl1b5NaGlbSkqw84w75xUCVafkr_M/sendDocument'
        }

    @classmethod
    def send_file(cls, filename):
        with open(filename, 'rb') as f:
            files = {'document': f}
            requests.post(cls.data_for_send_msg['url'], data=cls.data_for_send_msg, files=files)

        os.remove(os.path.join(folder, filename))  # Удаляем файл после отправки в мессенджере


class Robot:
    def __init__(self, text):
        self.text = text

    def start(self):
        parser = Parser(self.text)
        data_from_parsing = parser.all_data_from_hh()
        strategy = FilesCreator(Pdf, data_from_parsing)
        file_name = strategy.create_file()
        Telegram.send_file(file_name)


if __name__ == '__main__':
    andrew = Robot('Рэпер')
    andrew.start()

